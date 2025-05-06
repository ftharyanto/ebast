from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import QcFmRecord
from .forms import QcFmRecordForm
from django.shortcuts import render
import requests, openpyxl, datetime, os
import pandas as pd
from django.http import JsonResponse, HttpResponse
from core.models import Operator
from io import StringIO
from openpyxl.utils.dataframe import dataframe_to_rows
from django.views import View
from django.shortcuts import redirect

class QcFmRecordListView(ListView):
    model = QcFmRecord
    template_name = 'qcfm/qcfmrecord_list.html'
    context_object_name = 'qcfmrecords'

class QcFmRecordCreateView(CreateView):
    model = QcFmRecord
    form_class = QcFmRecordForm
    template_name = 'qcfm/qcfmrecord_form.html'
    success_url = reverse_lazy('qcfm:qcfmrecord_list')

class QcFmRecordUpdateView(UpdateView):
    model = QcFmRecord
    form_class = QcFmRecordForm
    template_name = 'qcfm/qcfmrecord_form.html'
    success_url = reverse_lazy('qcfm:qcfmrecord_list')

    def form_valid(self, form):
        return super().form_valid(form)

class QcFmRecordDeleteDirectView(View):
    def post(self, request, pk, *args, **kwargs):
        try:
            record = QcFmRecord.objects.get(pk=pk)
            record.delete()
            return redirect('qcfm:qcfmrecord_list')
        except QcFmRecord.DoesNotExist:
            return HttpResponse(status=404)


# Functions
def clean_fm_data(data, start_datetime='2025-03-12 00:00:00', end_datetime='2025-03-13 00:00:00'):
    text = data.decode('utf-8')

    lines = text.split('\n')
    processed_lines = []
    for i, line in enumerate(lines):
        line = '|'.join(part.strip() for part in line.split('|'))
        processed_lines.append(line)

    df = pd.DataFrame([x.split('|') for x in processed_lines[1:]], columns=processed_lines[0].split('|'))
    print(df)
    df['Datetime (UTC)'] = pd.to_datetime(df['Datetime (UTC)'], format='%Y-%m-%d %H:%M:%S')

    def select_data_by_datetime_range(df, start_datetime, end_datetime):
        mask = (df['Datetime (UTC)'] >= start_datetime) & (df['Datetime (UTC)'] <= end_datetime)
        return df.loc[mask]

    df_selected = select_data_by_datetime_range(df, start_datetime, end_datetime)

    # sort the df_selected by 'Datetime (UTC)'
    df_selected = df_selected.sort_values(by='Datetime (UTC)')

    # divide the 'Datetime (UTC)' column into 'Date' and 'OT (UTC)' columns, and put them in the first two columns, remove the 'Datetime (UTC)' column
    df_selected['Date'] = df_selected['Datetime (UTC)'].dt.date
    df_selected['OT (UTC)'] = df_selected['Datetime (UTC)'].dt.time
    df_selected = df_selected[['Date', 'OT (UTC)'] + [col for col in df_selected.columns if col != 'Datetime (UTC)']]
    df_selected = df_selected.reset_index(drop=True)

    # sort the columns to be 'Date', 'OT (UTC)', 'Lat', 'Long', 'Mag', 'D(Km)', 'Phase', 'RMS', 'Az.Gap', 'Region', but first turn the respective column names into the desired ones
    df_selected = df_selected.rename(columns={'D': 'D(Km)', 'Type M': 'TypeMag'})
    df_selected = df_selected[['Date', 'OT (UTC)', 'Lat', 'Long', 'Mag', 'TypeMag', 'D(Km)', 'S1', 'D1', 'R1', 'S2', 'D2', 'R2', 'Fit(%)', 'CLVD(%)']]
    df_selected = df_selected.reset_index(drop=True)

    # Check for duplicate columns
    df_selected = df_selected.loc[:, ~df_selected.columns.duplicated()]
    
    return df_selected

def fetch_data(request, start_datetime='2025-03-12 00:00:00', end_datetime='2025-03-13 00:00:00'):
    url = "http://202.90.198.41/qc_focal.txt"
    response = requests.get(url)

    if response.status_code == 200:
        data = clean_fm_data(response.content, start_datetime, end_datetime)
        csv_data = data.to_csv(index=False)
        table_data = data.insert(0, 'No', range(1, len(data) + 1))
        table_data = data.to_dict(orient='records')
        return JsonResponse({'csv': csv_data, 'table_data': table_data})
    else:
        return JsonResponse({'error': 'Failed to fetch data'}, status=500)

def get_nip(request, operator_id):
    try:
        operator = Operator.objects.get(id=operator_id)
        return JsonResponse({'nip': operator.NIP})
    except Operator.DoesNotExist:
        return JsonResponse({'error': 'Operator not found'}, status=404)

def prepare_workbook(record):
    file_path = os.path.join(os.path.dirname(__file__), 'static/qcfm/QC_FM.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'QC Records'

    tanggal = format_date_indonesian(record.qcfm_id[5:-3])
    hari = get_hari_indonesia(record.qcfm_id[5:-3])
    sheet['G2'] = ': ' + tanggal
    sheet['G3'] = ': ' + hari
    sheet['G4'] = f': {record.jam_pelaksanaan.strftime("%H:%M")} - selesai'
    sheet['G5'] = f': Kel. {record.kelompok}'

    qcfm_prev = pd.read_csv(StringIO(record.qcfm_prev))
    qcfm_prev['prev'] = f'Kel. {record.kel_sebelum}'
    rows_to_add = len(qcfm_prev)
    sheet.insert_rows(8, amount=rows_to_add * 2)
    qcfm_prev = dataframe_to_rows(qcfm_prev, index=False, header=False)

    qcfm = pd.read_csv(StringIO(record.qcfm))
    qcfm['QCFM'] = 'QCFM'
    qcfm = dataframe_to_rows(qcfm, index=False, header=False)

    for r_idx, row in enumerate(qcfm_prev, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx * 2 + 6, column=2, value=r_idx).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet.cell(row=r_idx * 2 + 6, column=c_idx + 2, value=value).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet.cell(row=r_idx * 2 + 6, column=c_idx + 2).fill = openpyxl.styles.PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
        sheet.merge_cells(start_row=r_idx * 2 + 6, start_column=2, end_row=r_idx * 2 + 7, end_column=2)

    for r_idx, row in enumerate(qcfm, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx * 2 + 7, column=c_idx + 2, value=value).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for r_idx in range(rows_to_add * 2):
        for c_idx in range(17):
            sheet.cell(row=r_idx + 8, column=c_idx + 2).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        sheet.row_dimensions[r_idx + 8].height = 15

    sheet.cell(row=8 + rows_to_add * 2 + 2, column=13, value=f'Jakarta, {tanggal}')
    sheet.row_dimensions[8 + rows_to_add * 2 + 2].height = 23.5
    sheet.row_dimensions[8 + rows_to_add * 2 + 3].height = 23.5
    sheet.cell(row=8 + rows_to_add * 2 + 8, column=13, value=record.operator.name).font = openpyxl.styles.Font(name='Calibri', underline='single', size=18, bold=True)
    sheet.row_dimensions[8 + rows_to_add * 2 + 8].height = 23.5
    sheet.cell(row=8 + rows_to_add * 2 + 9, column=13, value='NIP. ' + record.operator.NIP)

    return workbook

def export_to_excel(request, record_id):
    try:
        record = QcFmRecord.objects.get(id=record_id)
    except QcFmRecord.DoesNotExist:
        return HttpResponse(status=404)

    workbook = prepare_workbook(record)
    def simplify_qcfm_id(qcfm_id):
        import re
        return re.sub(r'-(\d)([DPSM])$', r'-\2', qcfm_id)
    simple_qcfm_id = simplify_qcfm_id(record.qcfm_id)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={simple_qcfm_id}.xlsx'
    workbook.save(response)
    return response

def export_to_pdf(request, record_id):
    try:
        record = QcFmRecord.objects.get(id=record_id)
    except QcFmRecord.DoesNotExist:
        return HttpResponse(status=404)

    workbook = prepare_workbook(record)
    def simplify_qcfm_id(qcfm_id):
        import re
        return re.sub(r'-(\d)([DPSM])$', r'-\2', qcfm_id)
    simple_qcfm_id = simplify_qcfm_id(record.qcfm_id)
    temp_xlsx = os.path.join(os.path.dirname(__file__), f'static/qcfm/{simple_qcfm_id}.xlsx')
    temp_pdf_dir = os.path.join(os.path.dirname(__file__), 'static/qcfm')
    temp_pdf = os.path.join(temp_pdf_dir, f'{simple_qcfm_id}.pdf')

    workbook.save(temp_xlsx)

    import subprocess
    try:
        command = ['soffice', '--headless', '--convert-to', 'pdf:calc_pdf_Export', temp_xlsx, '--outdir', temp_pdf_dir]
        subprocess.run(command, check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error converting {temp_xlsx} to PDF: {e}")
    finally:
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)

    with open(temp_pdf, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename={simple_qcfm_id}.pdf'

    if os.path.exists(temp_pdf):
        os.remove(temp_pdf)

    return response


def format_date_indonesian(date_string):
    """Formats a date string in YYYY-MM-DD format into Indonesian date format.

    Args:
    date_string: The date string in YYYY-MM-DD format.

    Returns:
    The formatted date string in Indonesian format (DD MMMM YYYY).
    """

    date_obj = datetime.datetime.strptime(date_string, "%Y-%m-%d")
    bulan_indonesia = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember"
    }
    formatted_date = f"{date_obj.day} {bulan_indonesia[date_obj.month]} {date_obj.year}"
    return formatted_date

def get_hari_indonesia(date_string):
    """
    Converts a date string (YYYY-MM-DD) to its corresponding Indonesian day name.
    Works reliably across operating systems.
    """

    import datetime
    date_obj = datetime.datetime.strptime(date_string, "%Y-%m-%d")
    day_of_week_num = date_obj.weekday()  # Monday is 0, Sunday is 6

    hari_indonesia_map = {
        0: "Senin",
        1: "Selasa",
        2: "Rabu",
        3: "Kamis",
        4: "Jumat",
        5: "Sabtu",
        6: "Minggu"
    }

    return hari_indonesia_map[day_of_week_num]