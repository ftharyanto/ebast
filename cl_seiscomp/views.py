from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.shortcuts import render
import requests, openpyxl, datetime, os
from django.http import JsonResponse, HttpResponse
from core.models import Operator
from io import StringIO
from django.views import View
from django.shortcuts import redirect
from .models import CsRecordModel, StationListModel
from django.contrib import messages
import csv
import plotly.graph_objects as go
import plotly.utils
import json

# Define time choices
WAKTU = (
    ('00:00 WIB', '00:00 WIB'),
    ('06:00 WIB', '06:00 WIB'),
    ('12:00 WIB', '12:00 WIB'),
    ('18:00 WIB', '18:00 WIB'),
)

# Define shift choices
SHIFT = (
    ('Pagi', 'Pagi'),
    ('Siang', 'Siang'),
    ('Malam', 'Malam'),
    ('Dini Hari', 'Dini Hari'),
)

##### Station List Views
class StationListView(ListView):
    model = StationListModel
    template_name = 'cl_seiscomp/station_list.html'
    context_object_name = 'station_list'

class StationCreateView(CreateView):
    model = StationListModel
    template_name = 'cl_seiscomp/sl_form.html'
    fields = '__all__'
    success_url = reverse_lazy('cl_seiscomp:station_list')

class StationUpdateView(UpdateView):
    model = StationListModel
    template_name = 'cl_seiscomp/sl_form.html'
    fields = '__all__'
    success_url = reverse_lazy('cl_seiscomp:station_list')

class StationDeleteView(DeleteView):
    model = StationListModel
    template_name = 'cl_seiscomp/sl_confirm_delete.html'
    success_url = reverse_lazy('cl_seiscomp:station_list')

class StationBulkCreateView(View):
    template_name = 'cl_seiscomp/sl_bulk_create.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        csv_file = request.FILES.get('csv_file')
        csv_data = request.POST.get('csv_data')

        if not csv_file and not csv_data:
            messages.error(request, 'Please upload a CSV file or provide CSV data')
            return redirect('cl_seiscomp:sl_bulk_create')

        if csv_file:
            if not csv_file.name.endswith('.csv'):
                messages.error(request, 'This is not a CSV file')
                return redirect('cl_seiscomp:sl_bulk_create')

            file_data = csv_file.read().decode('utf-8')
            csv_reader = csv.reader(StringIO(file_data), delimiter=',')
            header = next(csv_reader)  # Skip the header row
        else:
            csv_reader = csv.reader(StringIO(csv_data), delimiter=',')

        for row in csv_reader:
            StationListModel.objects.create(
                network=row[0],
                code=row[1],
                province=row[2],
                location=row[3],
                digitizer_type=row[4],
                UPT=row[5]
            )

        messages.success(request, 'Stations added successfully')
        return redirect('cl_seiscomp:station_list')


##### Stats View
class StatsView(View):
    template_name = 'cl_seiscomp/stats.html'

    def get(self, request):
        # Get date range from request parameters
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        selected_time = request.GET.get('time')
        
        # Set default date range to last 30 days if not specified
        if not start_date or not end_date:
            end_date = datetime.date.today()
            start_date = end_date - datetime.timedelta(days=30)
        else:
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Get records within the date range
        records = CsRecordModel.objects.filter(date__range=[start_date, end_date])
        
        # Filter by time if specified
        if selected_time and selected_time != 'All':
            records = records.filter(jam_pelaksanaan=selected_time)
        
        # Prepare data for charts
        dates = []
        gaps_data = []
        blanks_data = []
        spikes_data = []
        slmon_data = []
        
        # Calculate error frequencies per station
        station_errors = {}
        
        for record in records:
            dates.append(f"{record.date.strftime('%Y-%m-%d')} {record.jam_pelaksanaan}")
            gaps_data.append(record.count_gaps)
            blanks_data.append(record.count_blanks)
            spikes_data.append(record.count_spikes)
            slmon_data.append(record.slmon)
            
            # Process gaps
            if record.gaps:
                for station in record.gaps.splitlines():
                    station_errors[station] = station_errors.get(station, {'gaps': 0, 'spikes': 0, 'blanks': 0})
                    station_errors[station]['gaps'] += 1
            
            # Process spikes
            if record.spikes:
                for station in record.spikes.splitlines():
                    station_errors[station] = station_errors.get(station, {'gaps': 0, 'spikes': 0, 'blanks': 0})
                    station_errors[station]['spikes'] += 1
            
            # Process blanks
            if record.blanks:
                for station in record.blanks.splitlines():
                    station_errors[station] = station_errors.get(station, {'gaps': 0, 'spikes': 0, 'blanks': 0})
                    station_errors[station]['blanks'] += 1
        
        # Sort stations by total errors (descending)
        sorted_stations = sorted(station_errors.items(), 
                               key=lambda x: sum(x[1].values()), 
                               reverse=True)
        
        # Create Plotly figure for main chart
        fig = go.Figure()

        # Add traces
        fig.add_trace(go.Scatter(x=dates, y=gaps_data, name='Gaps', line=dict(color='rgb(75, 192, 192)')))
        fig.add_trace(go.Scatter(x=dates, y=blanks_data, name='Blanks', line=dict(color='rgb(255, 99, 132)')))
        fig.add_trace(go.Scatter(x=dates, y=spikes_data, name='Spikes', line=dict(color='rgb(54, 162, 235)')))
        fig.add_trace(go.Scatter(x=dates, y=slmon_data, name='SLMON', line=dict(color='rgb(255, 205, 86)')))

        # Update layout with enhanced styling
        fig.update_layout(
            title={
                'text': f'Station Statistics ({start_date.strftime("%b %Y")} - {end_date.strftime("%b %Y")})',
                'font': {'size': 24, 'color': '#333'},
                'x': 0.5,
                'xanchor': 'center'
            },
            xaxis_title={
                'text': 'Date and Time',
                'font': {'size': 16, 'color': '#666'}
            },
            yaxis_title={
                'text': 'Count',
                'font': {'size': 16, 'color': '#666'}
            },
            xaxis={
                'showgrid': True,
                'gridcolor': '#eee',
                'tickfont': {'size': 12},
                'tickangle': -45
            },
            yaxis={
                'showgrid': True,
                'gridcolor': '#eee',
                'tickfont': {'size': 12}
            },
            hovermode='x unified',
            hoverlabel={
                'font_size': 14,
                'font_family': 'Arial'
            },
            legend={
                'orientation': 'h',
                'yanchor': 'bottom',
                'y': 1.02,
                'xanchor': 'right',
                'x': 1,
                'font': {'size': 14},
                'title': {'font': {'size': 16}}
            },
            margin={'t': 80, 'b': 80, 'l': 80, 'r': 80},
            plot_bgcolor='#fff',
            paper_bgcolor='#fff',
            annotations=[
                {
                    'text': 'Click and drag to zoom, double-click to reset zoom',
                    'x': 1.1,
                    'y': -0.7,
                    'xref': 'paper',
                    'yref': 'paper',
                    'showarrow': False,
                    'font': {'size': 12, 'color': '#999'},
                    'bgcolor': 'rgba(255, 255, 255, 0.8)',
                    'align': 'right'
                }
            ]
        )

        # Create second figure for station error frequencies
        station_fig = go.Figure()
        
        # Create traces for each error type
        station_names = [station[0] for station in sorted_stations]
        gaps_freq = [station[1]['gaps'] for station in sorted_stations]
        spikes_freq = [station[1]['spikes'] for station in sorted_stations]
        blanks_freq = [station[1]['blanks'] for station in sorted_stations]
        
        station_fig.add_trace(go.Bar(
            x=station_names,
            y=gaps_freq,
            name='Gaps',
            marker_color='rgb(75, 192, 192)'
        ))
        station_fig.add_trace(go.Bar(
            x=station_names,
            y=spikes_freq,
            name='Spikes',
            marker_color='rgb(54, 162, 235)'
        ))
        station_fig.add_trace(go.Bar(
            x=station_names,
            y=blanks_freq,
            name='Blanks',
            marker_color='rgb(255, 99, 132)'
        ))
        
        # Update layout for station chart
        station_fig.update_layout(
            title={
                'text': f'Frequency of Errors per Station ({start_date.strftime("%b %Y")} - {end_date.strftime("%b %Y")})',
                'font': {'size': 24, 'color': '#333'},
                'x': 0.5,
                'xanchor': 'center'
            },
            xaxis_title={
                'text': 'Station',
                'font': {'size': 16, 'color': '#666'}
            },
            yaxis_title={
                'text': 'Count',
                'font': {'size': 16, 'color': '#666'}
            },
            barmode='stack',
            xaxis={
                'showgrid': True,
                'gridcolor': '#eee',
                'tickfont': {'size': 12},
                'tickangle': -45
            },
            yaxis={
                'showgrid': True,
                'gridcolor': '#eee',
                'tickfont': {'size': 12}
            },
            hovermode='x unified',
            hoverlabel={
                'font_size': 14,
                'font_family': 'Arial'
            },
            legend={
                'orientation': 'h',
                'yanchor': 'bottom',
                'y': 1.02,
                'xanchor': 'right',
                'x': 1,
                'font': {'size': 14},
                'title': {'font': {'size': 16}}
            },
            margin={'t': 80, 'b': 80, 'l': 80, 'r': 80},
            plot_bgcolor='#fff',
            paper_bgcolor='#fff',
            annotations=[
                {
                    'text': 'Click and drag to zoom, double-click to reset zoom',
                    'x': 1,
                    'y': -0.3,
                    'xref': 'paper',
                    'yref': 'paper',
                    'showarrow': False,
                    'font': {'size': 12, 'color': '#999'},
                    'bgcolor': 'rgba(255, 255, 255, 0.8)',
                    'align': 'right'
                }
            ]
        )
        
        # Convert figures to JSON
        plot_json = json.dumps(fig, cls=plotly.utils.PlotlyJSONEncoder)
        station_plot_json = json.dumps(station_fig, cls=plotly.utils.PlotlyJSONEncoder)

        context = {
            'plot_json': plot_json,
            'station_plot_json': station_plot_json,
            'start_date': start_date,
            'end_date': end_date,
            'times': WAKTU,
            'selected_time': selected_time or 'All'
        }
        return render(request, self.template_name, context)

##### Checklist Seiscomp View
from django.core.paginator import Paginator

class CsListView(ListView):
    model = CsRecordModel
    template_name = 'cl_seiscomp/cs_list.html'
    context_object_name = 'csrecords'
    paginate_by = 10
    ordering = ['-cs_id']

    def get_paginate_by(self, queryset):
        if self.request.GET.get('all') == '1':
            return None
        return self.paginate_by

    def get_queryset(self):
        return super().get_queryset().order_by('-cs_id')

class CsCreateView(CreateView):
    model = CsRecordModel
    template_name = 'cl_seiscomp/cs_form.html'
    fields = '__all__'
    success_url = reverse_lazy('cl_seiscomp:cs_list')

    def form_valid(self, form):
        form.instance.slmon_image = self.request.FILES.get('slmon_image')
        return super().form_valid(form)

class CsUpdateView(UpdateView):
    model = CsRecordModel
    template_name = 'cl_seiscomp/cs_form.html'
    fields = '__all__'
    success_url = reverse_lazy('cl_seiscomp:cs_list')

    def form_valid(self, form):
        if 'clear_image' in self.request.POST:
            form.instance.slmon_image = None
        else:
            if not self.request.FILES.get('slmon_image'):
                form.instance.slmon_image = self.get_object().slmon_image
            else:
                form.instance.slmon_image = self.request.FILES.get('slmon_image')
        return super().form_valid(form)

class CsDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        try:
            record = CsRecordModel.objects.get(pk=pk)
            record.delete()
            return redirect('cl_seiscomp:cs_list')
        except CsRecordModel.DoesNotExist:
            return HttpResponse(status=404)

def prepare_workbook(record):
    from qc.views import format_date_indonesian, get_hari_indonesia
    from datetime import timedelta
    from openpyxl.drawing.image import Image

    file_path = os.path.join(os.path.dirname(__file__), 'static/cl_seiscomp/cl_seiscomp.xlsx')
    workbook = openpyxl.load_workbook(file_path)

    # Prepare checklist_seiscomp sheet
    sheet = workbook['checklist_seiscomp']
    sheet.title = 'Checklist Seiscomp'

    tanggal = datetime.datetime.strptime(record.cs_id[3:-3], "%Y-%m-%d")
    if record.shift.upper() == 'MALAM':
        tanggal = date_range_to_string([tanggal, tanggal + timedelta(days=1)])
        sheet['R3'] = tanggal
    else:
        tanggal = record.cs_id[3:-3]
        sheet['R3'] = f'{get_hari_indonesia(tanggal)}, {format_date_indonesian(tanggal)}'

    sheet['A3'] = f'KELOMPOK: {record.kelompok}'
    sheet['A2'] = f'SHIFT {record.shift.upper()}'
    sheet['H266'] = f'{record.operator}'
    jam_pelaksanaan = f'JAM {record.jam_pelaksanaan}'
    sheet['D5'], sheet['P5'], sheet['H253'] = jam_pelaksanaan, jam_pelaksanaan, jam_pelaksanaan

    gaps = record.gaps.splitlines() if record.gaps else []
    spikes = record.spikes.splitlines() if record.spikes else []
    blanks = record.blanks.splitlines() if record.blanks else []

    for row in range(7, 269+1):
        cell_value = sheet.cell(row=row, column=2).value
        if cell_value in gaps:
            sheet.cell(row=row, column=4).value = 1
        if cell_value in spikes:
            sheet.cell(row=row, column=5).value = 1
        if cell_value in blanks:
            sheet.cell(row=row, column=6).value = 1

    for row in range(7, 249+1):
        cell_value = sheet.cell(row=row, column=9).value
        if cell_value in gaps:
            sheet.cell(row=row, column=16).value = 1
        if cell_value in spikes:
            sheet.cell(row=row, column=17).value = 1
        if cell_value in blanks:
            sheet.cell(row=row, column=18).value = 1

    # Prepare slmon sheet
    sheet = workbook['slmon']
    tanggal = datetime.datetime.strptime(record.cs_id[3:-3], "%Y-%m-%d")
    if record.shift.upper() == 'MALAM':
        tanggal = (tanggal + timedelta(days=1)).strftime('%Y-%m-%d')
    else:
        tanggal = tanggal.strftime('%Y-%m-%d')

    tanggal = format_date_indonesian(tanggal)
    sheet['A2'] = f'{tanggal}, pukul {record.jam_pelaksanaan}'
    sheet['M24'] = f'Jakarta, {tanggal}'
    sheet['C28'] = f'{record.operator}'

    if record.slmon_image:
        try:
            img = Image(record.slmon_image.path)
            img.anchor = 'B3'
            img.width = 8.6 * 96
            img.height = 4.14 * 96
            if len(sheet._images) == 0:
                sheet.add_image(img)
            else:
                sheet._images[0] = img
        except FileNotFoundError:
            sheet['B3'] = 'Image file not found'
    else:
        sheet['B3'] = 'No image'

    return workbook

def cs_export_excel(request, record_id):
    try:
        record = CsRecordModel.objects.get(id=record_id)
    except CsRecordModel.DoesNotExist:
        return HttpResponse(status=404)

    workbook = prepare_workbook(record)
    def simplify_cs_id(cs_id):
        import re
        return re.sub(r'-(\d)([DPSM])$', r'-\2', cs_id)
    simple_cs_id = simplify_cs_id(record.cs_id)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={simple_cs_id}.xlsx'
    workbook.save(response)
    return response

def cs_export_pdf(request, record_id):
    import subprocess

    try:
        record = CsRecordModel.objects.get(id=record_id)
    except CsRecordModel.DoesNotExist:
        return HttpResponse(status=404)

    workbook = prepare_workbook(record)
    def simplify_cs_id(cs_id):
        import re
        return re.sub(r'-(\d)([DPSM])$', r'-\2', cs_id)
    simple_cs_id = simplify_cs_id(record.cs_id)
    temp_xlsx = os.path.join(os.path.dirname(__file__), f'static/cl_seiscomp/{simple_cs_id}.xlsx')
    temp_pdf_dir = os.path.join(os.path.dirname(__file__), 'static/cl_seiscomp')
    temp_pdf = os.path.join(temp_pdf_dir, f'{simple_cs_id}.pdf')

    workbook.save(temp_xlsx)

    try:
        command = ['soffice', '--headless', '--convert-to', 'pdf:calc_pdf_Export', temp_xlsx, '--outdir', temp_pdf_dir]
        subprocess.run(command, check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error converting {temp_xlsx} to PDF: {e}")
    finally:
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)

    with open(temp_pdf, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename={simple_cs_id}.pdf'

    if os.path.exists(temp_pdf):
        os.remove(temp_pdf)
    return response

def date_range_to_string(date_range):
    import locale
    locale.setlocale(locale.LC_TIME, "id_ID.utf8")
    
    weekdays = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    date_strings = date_range
    start_date = date_strings[0].strftime('%d')
    end_date = date_strings[-1].strftime('%d')
    start_month = date_strings[0].strftime('%B')
    end_month = date_strings[-1].strftime('%B')
    start_year = date_strings[0].strftime('%Y')
    end_year = date_strings[-1].strftime('%Y')
    start_weekday = weekdays[date_strings[0].weekday()]
    end_weekday = weekdays[date_strings[-1].weekday()]
    if (start_year != end_year) and (start_month != end_month):
        return f"{start_weekday} - {end_weekday}, {start_date} {start_month} {start_year} - {end_date} {end_month} {end_year}"
    elif (start_year == end_year) and (start_month != end_month):
        return f"{start_weekday} - {end_weekday}, {start_date} {start_month} - {end_date} {end_month} {start_year}"
    else:
        return f"{start_weekday} - {end_weekday}, {start_date} - {end_date} {start_month} {start_year}"

# function to get waveform data (gaps, blanks)
def fetch_gaps_blanks(request):
    # open the data located in static/cl_seiscomp/checklist.txt
    file_path = os.path.join(os.path.dirname(__file__), 'static/cl_seiscomp/checklist.txt')
    with open(file_path, 'r') as f:
        data = f.read()
        data = data.split('\n\n')
        gaps = data[2]
        # skip the first line
        gaps = gaps.split('\n')[1:]
        gaps = '\n'.join(gaps)
        blanks = data[1]
        # skip the first line
        blanks = blanks.split('\n')[1:]
        blanks = '\n'.join(blanks)
        last_update = data[0].split(' ')[1:]
        last_update = ' '.join(last_update)

    return JsonResponse({'last_update': last_update, 'gaps': gaps, 'blanks': blanks})