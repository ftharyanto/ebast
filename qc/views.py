from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import QcRecord
from .forms import QcRecordForm
from django.shortcuts import render
import requests, openpyxl, datetime, os
import pandas as pd
from django.http import JsonResponse, HttpResponse
from core.models import Operator
from io import StringIO
from openpyxl.utils.dataframe import dataframe_to_rows
from django.views import View
from django.shortcuts import redirect
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json

class QcRecordListView(ListView):
    model = QcRecord
    template_name = 'qc/qcrecord_list.html'
    context_object_name = 'qcrecords'

class QcRecordCreateView(CreateView):
    model = QcRecord
    form_class = QcRecordForm
    template_name = 'qc/qcrecord_form.html'
    success_url = reverse_lazy('qc:qcrecord_list')

class QcRecordUpdateView(UpdateView):
    model = QcRecord
    form_class = QcRecordForm
    template_name = 'qc/qcrecord_form.html'
    success_url = reverse_lazy('qc:qcrecord_list')

    def form_valid(self, form):
        return super().form_valid(form)

class QcRecordDeleteDirectView(View):
    def post(self, request, pk, *args, **kwargs):
        try:
            record = QcRecord.objects.get(pk=pk)
            record.delete()
            return redirect('qc:qcrecord_list')
        except QcRecord.DoesNotExist:
            return HttpResponse(status=404)


# Functions
def clean_index3(data, start_datetime='2024-12-11 13:00:00', end_datetime='2024-12-11 19:00:00'):
    text = data.decode('utf-8')

    lines = text.split('\n')
    processed_lines = []
    for i, line in enumerate(lines):
        if i not in [0, 1, 3]:
            line = '|'.join(part.strip() for part in line.split('|'))
            processed_lines.append(line)

    df = pd.DataFrame([x.split('|') for x in processed_lines[1:]], columns=processed_lines[0].split('|'))
    df['Origin Time (GMT)'] = pd.to_datetime(df['Origin Time (GMT)'], format='%Y-%m-%d %H:%M:%S')

    def select_data_by_datetime_range(df, start_datetime, end_datetime):
        mask = (df['Origin Time (GMT)'] >= start_datetime) & (df['Origin Time (GMT)'] <= end_datetime)
        return df.loc[mask]

    df_selected = select_data_by_datetime_range(df, start_datetime, end_datetime)

    # sort the df_selected by 'Origin Time (GMT)'
    df_selected = df_selected.sort_values(by='Origin Time (GMT)')

    # divide the 'Origin Time (GMT)' column into 'Date' and 'OT (UTC)' columns, and put them in the first two columns, remove the 'Origin Time (GMT)' column
    df_selected['Date'] = df_selected['Origin Time (GMT)'].dt.date
    df_selected['OT (UTC)'] = df_selected['Origin Time (GMT)'].dt.time
    df_selected = df_selected[['Date', 'OT (UTC)'] + [col for col in df_selected.columns if col != 'Origin Time (GMT)']]
    df_selected = df_selected.reset_index(drop=True)

    # sort the columns to be 'Date', 'OT (UTC)', 'Lat', 'Long', 'Mag', 'D(Km)', 'Phase', 'RMS', 'Az.Gap', 'Region', but first turn the respective column names into the desired ones
    df_selected = df_selected.rename(columns={'Lon': 'Long', 'Depth': 'D(Km)', 'cntP': 'Phase', 'AZgap': 'Az. Gap', 'Remarks': 'Region'})
    df_selected = df_selected[['Date', 'OT (UTC)', 'Lat', 'Long', 'Mag', 'TypeMag', 'D(Km)', 'Phase', 'RMS', 'Az. Gap', 'Region']]
    df_selected = df_selected.reset_index(drop=True)

    # Check for duplicate columns
    df_selected = df_selected.loc[:, ~df_selected.columns.duplicated()]
    
    return df_selected

def fetch_data(request, start_datetime='2024-12-11 13:00:00', end_datetime='2024-12-11 19:00:00'):
    url = "http://202.90.198.41/index3.txt"
    response = requests.get(url)

    if response.status_code == 200:
        data = clean_index3(response.content, start_datetime, end_datetime)
        csv_data = data.to_csv(index=False)
        table_data = data.insert(0, 'No', range(1, len(data) + 1))
        table_data = data.to_dict(orient='records')
        return JsonResponse({'csv': csv_data, 'table_data': table_data})
    else:
        return JsonResponse({'error': 'Failed to fetch data'}, status=500)

def get_nip(request, operator_id):
    try:
        operator = Operator.objects.get(id=operator_id)
        return JsonResponse({'nip': operator.NIP})
    except Operator.DoesNotExist:
        return JsonResponse({'error': 'Operator not found'}, status=404)

@csrf_exempt
def save_nip(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            operator_id = data.get('operator_id')
            nip = data.get('nip')

            if not operator_id or not nip:
                return JsonResponse({'error': 'Operator ID and NIP are required'}, status=400)

            operator = Operator.objects.get(id=operator_id)
            operator.NIP = nip
            operator.save()

            return JsonResponse({'success': 'NIP saved successfully'})
        except Operator.DoesNotExist:
            return JsonResponse({'error': 'Operator not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)

def populate_sheet_with_record(sheet, record):
    """
    Populates the given sheet with data from the record.
    """
    tanggal = format_date_indonesian(record.qc_id[3:-2])
    hari = get_hari_indonesia(record.qc_id[3:-2])
    sheet['G2'] = ': ' + tanggal
    sheet['G3'] = ': ' + hari
    sheet['G4'] = f': {record.jam_pelaksanaan.strftime("%H:%M")} - selesai'
    sheet['G5'] = f': Kel. {record.kelompok}'
    sheet['B6'] = f'Event di Indonesia: {record.event_indonesia}'
    sheet['G6'] = f'Event di Luar Negeri: {record.event_luar}'

    qc_prev = pd.read_csv(StringIO(record.qc_prev))
    qc_prev['prev'] = f'Kel. {record.kel_sebelum}'
    rows_to_add = len(qc_prev)
    sheet.insert_rows(8, amount=rows_to_add * 2)
    qc_prev = dataframe_to_rows(qc_prev, index=False, header=False)

    qc = pd.read_csv(StringIO(record.qc))
    qc['QC'] = 'QC'
    qc = dataframe_to_rows(qc, index=False, header=False)

    for r_idx, row in enumerate(qc_prev, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx * 2 + 6, column=2, value=r_idx).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet.cell(row=r_idx * 2 + 6, column=c_idx + 2, value=value).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet.cell(row=r_idx * 2 + 6, column=c_idx + 2).fill = openpyxl.styles.PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
        sheet.merge_cells(start_row=r_idx * 2 + 6, start_column=2, end_row=r_idx * 2 + 7, end_column=2)

    for r_idx, row in enumerate(qc, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx * 2 + 7, column=c_idx + 2, value=value).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for r_idx in range(rows_to_add * 2):
        sheet.cell(row=r_idx + 8, column=13).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
        for c_idx in range(13):
            sheet.cell(row=r_idx + 8, column=c_idx + 2).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        sheet.row_dimensions[r_idx + 8].height = 15

    return rows_to_add, tanggal

def export_to_excel(request, record_id):
    try:
        record = QcRecord.objects.get(id=record_id)
    except QcRecord.DoesNotExist:
        return HttpResponse(status=404)

    file_path = os.path.join(os.path.dirname(__file__), 'static/qc/QC Seiscomp.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'QC Records'

    rows_to_add, tanggal = populate_sheet_with_record(sheet, record)

    sheet.cell(row=8 + rows_to_add * 2 + 2, column=13, value=f'Jakarta, {tanggal}')
    sheet.row_dimensions[8 + rows_to_add * 2 + 2].height = 23.5
    sheet.row_dimensions[8 + rows_to_add * 2 + 3].height = 23.5
    sheet.cell(row=8 + rows_to_add * 2 + 8, column=13, value=record.operator.name).font = openpyxl.styles.Font(name='Calibri', underline='single', size=18, bold=True)
    sheet.row_dimensions[8 + rows_to_add * 2 + 8].height = 23.5
    sheet.cell(row=8 + rows_to_add * 2 + 9, column=13, value='NIP. ' + record.operator.NIP)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=QC_{record.qc_id}.xlsx'
    workbook.save(response)

    return response

def export_to_pdf(request, record_id):
    try:
        record = QcRecord.objects.get(id=record_id)
    except QcRecord.DoesNotExist:
        return HttpResponse(status=404)

    file_path = os.path.join(os.path.dirname(__file__), 'static/qc/QC Seiscomp.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'QC Records'

    rows_to_add, tanggal = populate_sheet_with_record(sheet, record)

    sheet.cell(row=8 + rows_to_add * 2 + 2, column=13, value=f'Jakarta, {tanggal}')
    sheet.row_dimensions[8 + rows_to_add * 2 + 2].height = 23.5
    sheet.row_dimensions[8 + rows_to_add * 2 + 3].height = 23.5
    sheet.cell(row=8 + rows_to_add * 2 + 8, column=13, value=record.operator.name).font = openpyxl.styles.Font(name='Calibri', underline='single', size=18, bold=True)
    sheet.row_dimensions[8 + rows_to_add * 2 + 8].height = 23.5
    sheet.cell(row=8 + rows_to_add * 2 + 9, column=13, value='NIP. ' + record.operator.NIP)
    
    temp_xlsx = os.path.join(os.path.dirname(__file__), f'static/qc/{record.qc_id}.xlsx')
    workbook.save(temp_xlsx)
    temp_pdf_dir = os.path.join(os.path.dirname(__file__), 'static/qc')
    temp_pdf = os.path.join(temp_pdf_dir, f'{record.qc_id}.pdf')

    import subprocess
    try:
        command = ['soffice', '--headless', '--convert-to', 'pdf:calc_pdf_Export', temp_xlsx, '--outdir', temp_pdf_dir]
        subprocess.run(command, check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error converting {temp_xlsx} to PDF: {e}")
    finally:
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)

    with open(temp_pdf, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename={record.qc_id}.pdf'

    if os.path.exists(temp_pdf):
        os.remove(temp_pdf)

    return response

def format_date_indonesian(date_string):
    """Formats a date string in YYYY-MM-DD format into Indonesian date format.

    Args:
    date_string: The date string in YYYY-MM-DD format.

    Returns:
    The formatted date string in Indonesian format (DD MMMM YYYY).
    """

    date_obj = datetime.datetime.strptime(date_string, "%Y-%m-%d")
    bulan_indonesia = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember"
    }
    formatted_date = f"{date_obj.day} {bulan_indonesia[date_obj.month]} {date_obj.year}"
    return formatted_date

def get_hari_indonesia(date_string):
    """
    Converts a date string (YYYY-MM-DD) to its corresponding Indonesian day name.
    Works reliably across operating systems.
    """

    import datetime
    import calendar
    date_obj = datetime.datetime.strptime(date_string, "%Y-%m-%d")
    day_of_week_num = date_obj.weekday()  # Monday is 0, Sunday is 6

    hari_indonesia_map = {
        0: "Senin",
        1: "Selasa",
        2: "Rabu",
        3: "Kamis",
        4: "Jumat",
        5: "Sabtu",
        6: "Minggu"
    }

    return hari_indonesia_map[day_of_week_num]