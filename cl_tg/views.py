from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.shortcuts import render
import requests, openpyxl, datetime, os
from django.http import JsonResponse, HttpResponse
from core.models import Operator
from io import StringIO
from django.views import View
from django.shortcuts import redirect
from .models import CtgRecordModel, StationListModel
from django.contrib import messages
import csv

##### Station List Views
class StationListView(ListView):
    model = StationListModel
    template_name = 'cl_tg/station_list.html'
    context_object_name = 'station_list'

class StationCreateView(CreateView):
    model = StationListModel
    template_name = 'cl_tg/sl_form.html'
    fields = '__all__'
    success_url = reverse_lazy('cl_tg:station_list')

class StationUpdateView(UpdateView):
    model = StationListModel
    template_name = 'cl_tg/sl_form.html'
    fields = '__all__'
    success_url = reverse_lazy('cl_tg:station_list')

class StationBulkCreateView(View):
    template_name = 'cl_tg/sl_bulk_create.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        csv_file = request.FILES.get('csv_file')
        csv_data = request.POST.get('csv_data')

        if not csv_file and not csv_data:
            messages.error(request, 'Please upload a CSV file or provide CSV data')
            return redirect('cl_tg:sl_bulk_create')

        if csv_file:
            if not csv_file.name.endswith('.csv'):
                messages.error(request, 'This is not a CSV file')
                return redirect('cl_tg:sl_bulk_create')

            file_data = csv_file.read().decode('utf-8')
            csv_reader = csv.reader(StringIO(file_data), delimiter=',')
            header = next(csv_reader)  # Skip the header row
        else:
            csv_reader = csv.reader(StringIO(csv_data), delimiter=',')

        for row in csv_reader:
            StationListModel.objects.create(
                network=row[0],
                code=row[1],
                province=row[2],
                location=row[3],
                digitizer_type=row[4],
                UPT=row[5]
            )

        messages.success(request, 'Stations added successfully')
        return redirect('cl_tg:station_list')


##### Checklist Seiscomp View
class CsListView(ListView):
    model = CtgRecordModel
    template_name = 'cl_tg/cl_tg_list.html'
    context_object_name = 'csrecords'

class CsCreateView(CreateView):
    model = CtgRecordModel
    template_name = 'cl_tg/cl_tg_form.html'
    fields = '__all__'
    success_url = reverse_lazy('cl_tg:cl_tg_list')

class CsUpdateView(UpdateView):
    model = CtgRecordModel
    template_name = 'cl_tg/cl_tg_form.html'
    fields = '__all__'
    success_url = reverse_lazy('cl_tg:cl_tg_list')

class CsDeleteView(View):
    def post(self, request, pk, *args, **kwargs):
        try:
            record = CtgRecordModel.objects.get(pk=pk)
            record.delete()
            return redirect('cl_tg:cl_tg_list')
        except CtgRecordModel.DoesNotExist:
            return HttpResponse(status=404)

def cs_export_excel(request, record_id):
    from qc.views import format_date_indonesian, get_hari_indonesia
    from datetime import timedelta

    try:
        record = CtgRecordModel.objects.get(id=record_id)
    except CtgRecordModel.DoesNotExist:
        return HttpResponse(status=404)

    file_path = os.path.join(os.path.dirname(__file__), 'static/cl_tg/cl_tg.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['FORM_CHECKLIST']
    sheet.title = 'FORM_CHECKLIST'
    
    tanggal = datetime.datetime.strptime(record.cs_id[3:-2], "%Y-%m-%d")
    print(record.shift)
    if record.shift.upper() == 'MALAM':
        tanggal = date_range_to_string([tanggal, tanggal + timedelta(days=1)])
        sheet['L3'] = tanggal
    else:
        tanggal = record.cs_id[3:-2]
        sheet['L3'] = f'{get_hari_indonesia(tanggal)}, {format_date_indonesian(tanggal)}'
        
    sheet['C3'] = f'{record.kelompok}'
    sheet['C4'] = f'{record.shift.upper()}'
    sheet['H266'] = f'{record.operator}'
    sheet['L4'] = f'{record.jam_pelaksanaan}'

    if record.gaps:
        gaps = record.gaps.splitlines()
    else:
        gaps = []

    if record.spikes:
        spikes = record.spikes.splitlines()
    else:
        spikes = []

    if record.blanks:
        blanks = record.blanks.splitlines()
    else:
        blanks = []

    # for cells B7:B269
    for row in range(12, 269+1):  # Iterate through rows 7 to 269
        cell_value = sheet.cell(row=row, column=2).value  # Get the cell value in column B

        if cell_value in gaps:
            sheet.cell(row=row, column=4).value = 1
        if cell_value in spikes:
            sheet.cell(row=row, column=5).value = 1
        if cell_value in blanks:
            sheet.cell(row=row, column=6).value = 1

    # for cells I7:I250
    for row in range(7, 250+1):  # Iterate through rows 7 to 250
        cell_value = sheet.cell(row=row, column=9).value  # Get the cell value in column B

        if cell_value in gaps:
            sheet.cell(row=row, column=16).value = 1
        if cell_value in spikes:
            sheet.cell(row=row, column=17).value = 1
        if cell_value in blanks:
            sheet.cell(row=row, column=18).value = 1

    # sheet of slmon
    from openpyxl.drawing.image import Image

    sheet = workbook['slmon']
    tanggal = format_date_indonesian(record.cs_id[3:-2])
    sheet['A2'] = f'{tanggal}, pukul {record.jam_pelaksanaan}'
    sheet['M24'] = f'Jakarta, {tanggal}'
    sheet['C28'] = f'{record.operator}'

    if record.slmon_image:
        img = Image(record.slmon_image.path)
        if len(sheet._images) == 0:
            # Initalise the `ref` data, do sheet.add_image(...)   
            img.anchor='B3'
            sheet.add_image(img)
            # set the size of the image in inches
            sheet._images[0].width = 8.6 * 96  # 96 DPI is the default resolution
            sheet._images[0].height = 4.14 * 96  # 96 DPI is the default resolution
            
        elif len(sheet._images) == 1:
            # Replace the first image do **only** the following:
            sheet._images[0] = img
            # Update the default anchor `A1` to your needs
            sheet._images[0].anchor='B3'
            # set the size of the image in inches
            sheet._images[0].width = 8.6 * 96  # 96 DPI is the default resolution
            sheet._images[0].height = 4.14 * 96  # 96 DPI is the default resolution
        else:
            raise(ValueError, "Found more than 1 Image!")
    else:
        sheet['B3'] = 'No image'

    # Save the workbook to a BytesIO object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={record.cs_id}.xlsx'
    workbook.save(response)

    return response

def cs_export_pdf(request, record_id):
    from qc.views import format_date_indonesian, get_hari_indonesia
    from datetime import timedelta

    try:
        record = CtgRecordModel.objects.get(id=record_id)
    except CtgRecordModel.DoesNotExist:
        return HttpResponse(status=404)

    file_path = os.path.join(os.path.dirname(__file__), 'static/cl_tg/cl_tg.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['checklist_seiscomp']
    sheet.title = 'Checklist Seiscomp'
    
    tanggal = datetime.datetime.strptime(record.cs_id[3:-2], "%Y-%m-%d")
    print(record.shift)
    if record.shift.upper() == 'MALAM':
        tanggal = date_range_to_string([tanggal, tanggal + timedelta(days=1)])
        sheet['R3'] = tanggal
    else:
        tanggal = record.cs_id[3:-2]
        sheet['R3'] = f'{get_hari_indonesia(tanggal)}, {format_date_indonesian(tanggal)}'
        
    sheet['A3'] = f'KELOMPOK: {record.kelompok}'
    sheet['A2'] = f'SHIFT {record.shift.upper()}'
    sheet['H266'] = f'{record.operator}'
    jam_pelaksanaan = f'JAM {record.jam_pelaksanaan}'
    sheet['D5'], sheet['P5'], sheet['H253'] = jam_pelaksanaan, jam_pelaksanaan, jam_pelaksanaan


    if record.gaps:
        gaps = record.gaps.splitlines()
    else:
        gaps = []

    if record.spikes:
        spikes = record.spikes.splitlines()
    else:
        spikes = []

    if record.blanks:
        blanks = record.blanks.splitlines()
    else:
        blanks = []

    # for cells B7:B109
    for row in range(7, 109+1):  # Iterate through rows 7 to 109
        cell_value = sheet.cell(row=row, column=2).value  # Get the cell value in column B

        if cell_value in gaps:
            sheet.cell(row=row, column=5).value = 1
        if cell_value in spikes:
            sheet.cell(row=row, column=6).value = 1
        if cell_value in blanks:
            sheet.cell(row=row, column=7).value = 1

    # for cells I7:I250
    for row in range(7, 250+1):  # Iterate through rows 7 to 250
        cell_value = sheet.cell(row=row, column=9).value  # Get the cell value in column B

        if cell_value in gaps:
            sheet.cell(row=row, column=16).value = 1
        if cell_value in spikes:
            sheet.cell(row=row, column=17).value = 1
        if cell_value in blanks:
            sheet.cell(row=row, column=18).value = 1

    # sheet of slmon
    from openpyxl.drawing.image import Image

    sheet = workbook['slmon']
    tanggal = format_date_indonesian(record.cs_id[3:-2])
    sheet['A2'] = f'{tanggal}, pukul {record.jam_pelaksanaan}'
    sheet['M24'] = f'Jakarta, {tanggal}'
    sheet['C28'] = f'{record.operator}'

    if record.slmon_image:
        img = Image(record.slmon_image.path)
        if len(sheet._images) == 0:
            # Initalise the `ref` data, do sheet.add_image(...)   
            img.anchor='B3'
            sheet.add_image(img)
            # set the size of the image in inches
            sheet._images[0].width = 8.6 * 96  # 96 DPI is the default resolution
            sheet._images[0].height = 4.14 * 96  # 96 DPI is the default resolution
            
        elif len(sheet._images) == 1:
            # Replace the first image do **only** the following:
            sheet._images[0] = img
            # Update the default anchor `A1` to your needs
            sheet._images[0].anchor='B3'
            # set the size of the image in inches
            sheet._images[0].width = 8.6 * 96  # 96 DPI is the default resolution
            sheet._images[0].height = 4.14 * 96  # 96 DPI is the default resolution
        else:
            raise(ValueError, "Found more than 1 Image!")
    else:
        sheet['B3'] = 'No image'

    # temporarily save the workbook to a file
    temp_xlsx = os.path.join(os.path.dirname(__file__), f'static/cl_tg/{record.cs_id}.xlsx')
    workbook.save(temp_xlsx)
    temp_pdf_dir = os.path.join(os.path.dirname(__file__), 'static/cl_tg')
    temp_pdf = os.path.join(temp_pdf_dir, f'{record.cs_id}.pdf')

    import subprocess
    try:
        command = ['soffice', '--headless', '--convert-to', 'pdf:impress_pdf_Export', temp_xlsx, '--outdir', temp_pdf_dir]
        subprocess.run(command, check=True)
        print(temp_xlsx)
    except subprocess.CalledProcessError as e:
        print(f"Error converting {temp_xlsx} to PDF: {e}")
    finally:
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)

    # Read the generated PDF file and return it in the response
    with open(temp_pdf, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename={record.cs_id}.pdf'

    if os.path.exists(temp_pdf):
        os.remove(temp_pdf)
    return response


def date_range_to_string(date_range):
    import locale
    locale.setlocale(locale.LC_TIME, "id_ID.utf8")
    
    weekdays = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    date_strings = date_range
    start_date = date_strings[0].strftime('%d')
    end_date = date_strings[-1].strftime('%d')
    start_month = date_strings[0].strftime('%B')
    end_month = date_strings[-1].strftime('%B')
    start_year = date_strings[0].strftime('%Y')
    end_year = date_strings[-1].strftime('%Y')
    start_weekday = weekdays[date_strings[0].weekday()]
    end_weekday = weekdays[date_strings[-1].weekday()]
    if (start_year != end_year) and (start_month != end_month):
        return f"{start_weekday} - {end_weekday}, {start_date} {start_month} {start_year} - {end_date} {end_month} {end_year}"
    elif (start_year == end_year) and (start_month != end_month):
        return f"{start_weekday} - {end_weekday}, {start_date} {start_month} - {end_date} {end_month} {start_year}"
    else:
        return f"{start_weekday} - {end_weekday}, {start_date} - {end_date} {start_month} {start_year}"

# function to get waveform data (gaps, blanks)
def fetch_gaps_blanks(request):
    # open the data located in static/cl_tg/checklist.txt
    file_path = os.path.join(os.path.dirname(__file__), 'static/cl_tg/checklist.txt')
    with open(file_path, 'r') as f:
        data = f.read()
        data = data.split('\n\n')
        gaps = data[2]
        # skip the first line
        gaps = gaps.split('\n')[1:]
        gaps = '\n'.join(gaps)
        blanks = data[1]
        # skip the first line
        blanks = blanks.split('\n')[1:]
        blanks = '\n'.join(blanks)
        last_update = data[0].split(' ')[1:]
        last_update = ' '.join(last_update)

    return JsonResponse({'last_update': last_update, 'gaps': gaps, 'blanks': blanks})