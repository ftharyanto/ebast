import openpyxl

def generate_excel(filename, response, metadata, data):
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.worksheets[0]

    # constants
    # nongaransi1 = [ws.cell(row=i,column=2).value for i in range(8, 267 +1)]
    # nongaransi2 = [ws.cell(row=i,column=10).value for i in range(8, 149 +1)]
    # garansi = [ws.cell(row=i,column=10).value for i in range(154, 246 +1)]
    stasiun1 = [ws.cell(row=i,column=2).value for i in range(7, 269 + 1)]
    stasiun2 = [ws.cell(row=i,column=9).value for i in range(7, 250 + 1)]


    def loop_cells(type, station_code, row, column):
        for i in range(len(station_code)):
            for j in range(len(type)):
                if type[j] == station_code[i]:
                    ws.cell(row=i + row, column=column).value = 1


    def clear_cells(cell_range: str, ws):
        for row in ws[cell_range]:
            for cell in row:
                cell.value = None


    def clear_contents():
        clear_cells('D7:F269', ws)
        clear_cells('P7:R250', ws)

    def update_metadata(metadata):
        ws['A3'] = f"KELOMPOK : {metadata['kelompok']}"

        ws['H266'] = metadata['operator']
        ws['R3'] = metadata['tanggal']
        ws['A2'] = metadata['shift']
        ws['D5'] = ws['P5'] = ws['H253'] = f"JAM {metadata['jam']}"


    def update_data(gaps, spikes, blanks):
        clear_cells('D7:F269', ws)
        clear_cells('P7:R250', ws)

        # nongaransi1 loop
        # loop_cells(gaps, nongaransi1, 8, 5)
        # loop_cells(spikes, nongaransi1, 8, 6)
        # loop_cells(blanks, nongaransi1, 8, 7)

        # nongaransi2 loop
        # loop_cells(gaps, nongaransi2, 8, 20)
        # loop_cells(spikes, nongaransi2, 8, 21)
        # loop_cells(blanks, nongaransi2, 8, 22)

        # garansi loop
        #loop_cells(gaps, garansi, 154, 20)
        #loop_cells(spikes, garansi, 154, 21)
        #loop_cells(blanks, garansi, 154, 22)
        
        # stasiun1
        loop_cells(gaps, stasiun1, 7, 4)
        loop_cells(spikes, stasiun1, 7, 5)
        loop_cells(blanks, stasiun1, 7, 6)
        
        # stasiun2
        loop_cells(gaps, stasiun2, 7, 16)
        loop_cells(spikes, stasiun2, 7, 17)
        loop_cells(blanks, stasiun2, 7, 18)
        
    update_data(data['gaps'], data['spikes'], data['blanks'])
    update_metadata(metadata)

    wb.save(response)
