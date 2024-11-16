from django.http import HttpResponse
from django.shortcuts import (get_object_or_404,
                              render,
                              HttpResponseRedirect)
import openpyxl
from django.urls import reverse

# relative import of forms
from .models import ChecklistSeiscompModel, OperatorModel, StationListModel
from .forms import InputForm, OperatorForm, StationListForm


def create_view(request):
    # dictionary for initial data with
    # field names as keys
    context = {}
    context['dataset'] = ChecklistSeiscompModel.objects.all().order_by('-id')[:2]

    # add the dictionary during initialization
    input_form = InputForm(request.POST or None)
    if input_form.is_valid():
        input_form.save()
        # Resetting the form after it has been submitted.
        input_form = InputForm()

    context['input_form'] = input_form
    return render(request, "create_view.html", context)


def list_view(request):
    # dictionary for initial data with
    # field names as keys
    context = {}

    # add the dictionary during initialization
    context["dataset"] = ChecklistSeiscompModel.objects.all().order_by('-id')

    return render(request, "list_view.html", context)


def update_view(request, id):
    # dictionary for initial data with
    # field names as keys
    context = {}

    # fetch the object related to passed id
    obj = get_object_or_404(ChecklistSeiscompModel, id=id)

    # pass the object as instance in form
    form = OperatorForm(request.POST or None, instance=obj)

    # save the data from the form and
    # redirect to detail_view
    if form.is_valid():
        form.save()
        return HttpResponseRedirect("/"+id)

    # add form dictionary to context
    context["form"] = form

    return render(request, "update_view.html", context)


def operator_delete(request, id):
    ob = OperatorModel.objects.get(id=id)
    ob.delete()
    return HttpResponseRedirect(reverse('ChecklistSeiscomp:operator_view'))


def data_delete(request, id):
    ob = ChecklistSeiscompModel.objects.get(id=id)
    ob.delete()
    return HttpResponseRedirect(reverse('ChecklistSeiscomp:list_view'))


def operator_update(request, id):
    ob = OperatorModel.objects.get(id=id)
    ob.name = request.POST.get('field1')
    ob.save()
    return HttpResponseRedirect(reverse('ChecklistSeiscomp:operator_view'))


def operator_view(request):
    # dictionary for initial data with
    # field names as keys
    context = {}

    # add the dictionary during initialization
    context["operators"] = OperatorModel.objects.all().order_by('name')

    add_operator_form = OperatorForm(request.POST or None)
    if add_operator_form.is_valid():
        add_operator_form.save()

        # Resetting the form after it has been submitted.
        add_operator_form = OperatorForm()

    context['add_operator_form'] = add_operator_form

    return render(request, "operator_view.html", context)


def edit_operator_name(request, id):
    # dictionary for initial data with
    # field names as keys
    context = {}

    # fetch the object related to passed id
    obj = get_object_or_404(OperatorModel, id=id)

    # pass the object as instance in form
    form = OperatorForm(request.POST or None, instance=obj)

    # save the data from the form and
    # redirect to detail_view
    if form.is_valid():
        form.save()
        return HttpResponseRedirect("/"+id)

    # add form dictionary to context
    context["form"] = form

    return render(request, "operator_view.html", context)

def send_to_gcp_ss(request):
    import requests, ast

    url = 'https://script.google.com/macros/s/AKfycbxPVOb8TVUbu8q8c23d4jEJdTDEJZLp9me8Z7rkv0pbQ0m7jcNhcC6jwgt2bNfEe_vTmA/exec'
    data = ChecklistSeiscompModel.objects.all().order_by('-tanggal')[:2]

    data1 = {}
    if data[1].gaps:
        data1['gaps'] = ast.literal_eval(data[1].gaps)
    else:
        data1['gaps'] = []

    if data[1].spikes:
        data1['spikes'] = ast.literal_eval(data[1].spikes)
    else:
        data1['spikes'] = []

    if data[1].blanks:
        data1['blanks'] = ast.literal_eval(data[1].blanks)
    else:
        data1['blanks'] = []    

    data2 = {}
    if data[0].gaps:
        data2['gaps'] = ast.literal_eval(data[0].gaps)
    else:
        data2['gaps'] = []    

    if data[0].spikes:
        data2['spikes'] = ast.literal_eval(data[0].spikes)
    else:
        data2['spikes'] = [] 

    if data[0].blanks:
        data2['blanks'] = ast.literal_eval(data[0].blanks)
    else:
        data2['blanks'] = [] 

    all_data = {
        'kelompok': data[0].kelompok,
        'operator1': data[0].operator,
        'operator2': data[1].operator,
        'tanggal': date_range_to_string([data[1].tanggal, data[0].tanggal]),
        'data1': {
            'gaps': data1['gaps'],
            'spikes': data1['spikes'],
            'blanks': data1['blanks'],
        },
        'data2': {
            'gaps': data2['gaps'],
            'spikes': data2['spikes'],
            'blanks': data2['blanks'],
        },
    }
    
    response = requests.post(url, json=all_data)
    print(response.text)
    print(all_data)
    return HttpResponseRedirect("/checklist-seiscomp/create_view")

def format_date_indonesian(d):
    from datetime import date
    # Define Indonesian names for days and months
    days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    months = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
              'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']

    # Get the day of the week, day of the month, month, and year from the date
    day_of_week = days[d.weekday()]
    day = d.day
    month = months[d.month - 1]
    year = d.year

    # Return the formatted date
    return f'{day_of_week}, {day} {month} {year}'

# def export_excel_instant(request):
#     """This function is used to export excel file containing last 2 records"""

#     from django.conf import settings
#     from .output_generator import generate_excel
#     import ast


#     data = ChecklistSeiscompModel.objects.all().order_by('-tanggal')[:2]
#     metadata = {'kelompok': data[0].kelompok,
#             'operator1': data[0].operator,
#             'operator2': data[1].operator,
#             'tanggal': date_range_to_string([data[1].tanggal, data[0].tanggal])}
    
#     data1 = {}
#     if data[1].gaps:
#         data1['gaps'] = ast.literal_eval(data[1].gaps)
#     else:
#         data1['gaps'] = []

#     if data[1].spikes:
#         data1['spikes'] = ast.literal_eval(data[1].spikes)
#     else:
#         data1['spikes'] = []

#     if data[1].blanks:
#         data1['blanks'] = ast.literal_eval(data[1].blanks)
#     else:
#         data1['blanks'] = []    

#     data2 = {}
#     if data[0].gaps:
#         data2['gaps'] = ast.literal_eval(data[0].gaps)
#     else:
#         data2['gaps'] = []    

#     if data[0].spikes:
#         data2['spikes'] = ast.literal_eval(data[0].spikes)
#     else:
#         data2['spikes'] = [] 

#     if data[0].blanks:
#         data2['blanks'] = ast.literal_eval(data[0].blanks)
#     else:
#         data2['blanks'] = [] 

#     # Get the path of the Excel file in static folder
#     file_path = str(settings.STATIC_ROOT) + '/ChecklistSeiscomp/template.xlsx'

#     # Save the workbook to a byte stream
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     generate_excel(filename=file_path,
#                    response=response,
#                    metadata=metadata,
#                    data1=data1,
#                    data2=data2)
    
#     # Set the file name and attachment header
#     response['Content-Disposition'] = 'attachment; filename="data.xlsx"'
#     # Return the response
#     return response


def export_excel(request, id):
    """This function is used to export excel file containing last 2 records"""

    from django.conf import settings
    from .output_generator import generate_excel
    import ast

    dataset = ChecklistSeiscompModel.objects.get(id=id)
    shift = ''
    if dataset.jam == '12:00 WIB':
        shift = 'SHIFT PAGI'
    elif dataset.jam == '18:00 WIB':
        shift = 'SHIFT SIANG'    
    else:
        shift = 'SHIFT MALAM'

    metadata = {'kelompok': dataset.kelompok,
            'operator': dataset.operator,
            'tanggal': format_date_indonesian(dataset.tanggal),
            'shift': shift,
            'jam': dataset.jam,
            }
    
    data = {}
    if dataset.gaps:
        data['gaps'] = ast.literal_eval(dataset.gaps)
    else:
        data['gaps'] = []

    if dataset.spikes:
        data['spikes'] = ast.literal_eval(dataset.spikes)
    else:
        data['spikes'] = []

    if dataset.blanks:
        data['blanks'] = ast.literal_eval(dataset.blanks)
    else:
        data['blanks'] = []   

    # Get the path of the Excel file in static folder
    file_path = str(settings.STATIC_ROOT) + '/ChecklistSeiscomp/template.xlsx'

    # Save the workbook to a byte stream
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    generate_excel(filename=file_path,
                   response=response,
                   metadata=metadata,
                   data=data,
                   )
    
    # Set the file name and attachment header
    response['Content-Disposition'] = f'attachment; filename="CS_{dataset.tanggal}_{shift.split()[1]}.xlsx"'
    # Return the response
    return response

def download_file(real_file_id):
    import io

    import google.auth
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    from googleapiclient.http import MediaIoBaseDownload
    from secret_settings import GSHEETS_CREDS
    """Downloads a file
    Args:
        real_file_id: ID of the file to download
    Returns : IO object with location.

    Load pre-authorized user credentials from the environment.
    TODO(developer) - See https://developers.google.com/identity
    for guides on implementing OAuth2 for the application.
    """
    creds = GSHEETS_CREDS

    try:
        # create drive api client
        service = build('drive', 'v3', credentials=creds)

        file_id = real_file_id

        # pylint: disable=maybe-no-member
        request = service.files().get_media(fileId=file_id)
        file = io.BytesIO()
        downloader = MediaIoBaseDownload(file, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            print(F'Download {int(status.progress() * 100)}.')

    except HttpError as error:
        print(F'An error occurred: {error}')
        file = None

    return file.getvalue()

def export_pdf_instant(request):
    # from django.conf import settings
    # from io import BytesIO
    # from reportlab.pdfgen import canvas
    # from reportlab.lib.pagesizes import letter
    # from openpyxl import load_workbook

    # def convert_to_pdf(file_path):
    #     wb = load_workbook(filename=file_path)
    #     ws = wb.active

    #     buffer = BytesIO()
    #     c = canvas.Canvas(buffer, pagesize=letter)

    #     for row in ws.iter_rows():
    #         for cell in row:
    #             c.drawString(100, 100, str(cell.value))

    #     c.save()

    #     buffer.seek(0)
    #     return buffer
    
    # static_folder = str(settings.STATIC_ROOT) + '/ChecklistSeiscomp/'
    # file_path = static_folder + 'template.xlsx'

    # buffer = convert_to_pdf(file_path)

    # response = HttpResponse(buffer.read(), content_type='application/pdf')
    # response['Content-Disposition'] = 'attachment; filename="file.pdf"'

    # return response

    # import gspread, sys, requests
    # print(sys.path)
    # from oauth2client.service_account import ServiceAccountCredentials
    # import secret_settings

    # creds = secret_settings.GSHEETS_CREDS

    # scope = [
    # 'https://www.googleapis.com/auth/spreadsheets',
    # 'https://www.googleapis.com/auth/drive'
    # ]

    # creds = gspread.oauth_from_dict(creds)
    # client=gspread.authorize(creds)

    # ssid = '1o4Bki_ofmU4WYMZbQlUvoY-JNBWvzeuyofxRM2ZFT6I'
    # gid = '169956573'
    # spreadsheet = client.open(ssid)
    # url = f'https://docs.google.com/spreadsheets/export?format=pdf&id={spreadsheet.id}&gid={gid}&portrait=true&size=Folio'
    # headers = {'Authorization': 'Bearer ' + creds.create_delegated("").get_access_token().access_token}
    # res = requests.get(url, headers=headers)
    # with open("Ceklis.pdf", 'wb') as f:
    #     f.write(res.content)

    # return res
    import requests

    url = 'https://script.google.com/macros/s/AKfycbxPVOb8TVUbu8q8c23d4jEJdTDEJZLp9me8Z7rkv0pbQ0m7jcNhcC6jwgt2bNfEe_vTmA/exec'
    response = requests.get(url)
    
    if response == 200:
        response['Content-Disposition'] = 'attachment; filename=fajar.pdf'
        return response
    else:
        return HttpResponseRedirect("/checklist-seiscomp/create_view")




def date_range_to_string(date_range):
    import locale
    
    weekdays = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    date_strings = date_range
    start_date = date_strings[0].strftime('%d')
    end_date = date_strings[-1].strftime('%d')
    locale.setlocale(locale.LC_TIME, "id_ID.utf8")
    start_month = date_strings[0].strftime('%B')
    end_month = date_strings[-1].strftime('%B')
    start_year = date_strings[0].strftime('%Y')
    end_year = date_strings[-1].strftime('%Y')
    start_weekday = weekdays[date_strings[0].weekday()]
    end_weekday = weekdays[date_strings[-1].weekday()]
    if (start_year != end_year) and (start_month != end_month):
        return f"{start_weekday} - {end_weekday}, {start_date} {start_month} {start_year} - {end_date} {end_month} {end_year}"
    elif (start_year == end_year) and (start_month != end_month):
        return f"{start_weekday} - {end_weekday}, {start_date} {start_month} - {end_date} {end_month} {start_year}"
    else:
        return f"{start_weekday} - {end_weekday}, {start_date} - {end_date} {start_month} {start_year}"

def station_list_view(request):
    # dictionary for initial data with
    context = {}

    # get count of garansi and nongaransi
    garansi_count = StationListModel.objects.filter(tipe='garansi').values_list('tipe', flat=True).count()
    nongaransi_count = StationListModel.objects.filter(tipe='nongaransi').values_list('tipe', flat=True).count()

    # add the dictionary during initialization
    context["station_list"] = StationListModel.objects.all()

    add_station_list_form = StationListForm(request.POST or None)
    if add_station_list_form.is_valid():
        add_station_list_form.save()

        # Resetting the form after it has been submitted.
        add_station_list_form = StationListForm()

    context['add_station_list_form'] = add_station_list_form
    context['garansi_count'] = garansi_count
    context['nongaransi_count'] = nongaransi_count

    return render(request, "station_list_view.html", context)        

def station_list_delete(request, id):
    ob = StationListModel.objects.get(id=id)
    ob.delete()
    return HttpResponseRedirect("/checklist-seiscomp/station_list_view")


def statistic_view(request):
    from django.db.models import Q
    from plotly.offline import plot
    import plotly.graph_objs as go
    from datetime import datetime
    import ast

    station_count = StationListModel.objects.count()
    last_data = ChecklistSeiscompModel.objects.all().last()

    # define x-axis as date time of 12:00 WIB
    x_data = list(ChecklistSeiscompModel.objects.filter(Q(jam='12:00 WIB')).values_list('tanggal', flat=True))
    x_datetime = []
    for x in x_data:
        x = datetime.strptime(x.strftime('%Y-%m-%d')+' 12:00', '%Y-%m-%d %H:%M')
        x_datetime.append(x)

    # graphic of visual monitoring vs slmon 12:00 WIB
    y_slmon_12 = list(ChecklistSeiscompModel.objects.filter(Q(jam='12:00 WIB')).values_list('slmon', flat=True))
    y_blanks_12 = list(ChecklistSeiscompModel.objects.filter(Q(jam='12:00 WIB')).values_list('blanks', flat=True))
    
    y_blanks_12_len = []
    for data in y_blanks_12:
        if data:
            y = len(ast.literal_eval(data))
            y_blanks_12_len.append(y)
        else:
            y_blanks_12_len.append(0)


    layout12=go.Layout(title="Grafik Slmon vs Visual Monitoring Pukul 12:00 WIB", title_x=0.5, xaxis={'title':'Waktu'}, yaxis={'title':'Jumlah'})

    fig12 = go.Figure(data=[
            go.Line(x=x_datetime, y=y_slmon_12,
                name='slmon 12:00 WIB',
                opacity=0.8, marker_color='red'),
            go.Line(x=x_datetime, y=y_blanks_12_len,
                name='blanks 12:00 WIB',
                opacity=0.8, marker_color='blue'),
            ],
            layout=layout12)
    
    jam12 = plot(fig12,
               output_type='div', include_plotlyjs=False, show_link=False
               )
    
    # # define x-axis as date time of 12:00 WIB
    # x_data = list(ChecklistSeiscompModel.objects.filter(Q(jam='12:00 WIB')).values_list('tanggal', flat=True))
    # x_datetime = []
    # for x in x_data:
    #     x = datetime.strptime(x.strftime('%Y-%m-%d')+' 00:00', '%Y-%m-%d %H:%M')
    #     x_datetime.append(x)

    # # graphic of visual monitoring vs slmon 00:00 WIB
    # y_slmon_00 = list(ChecklistSeiscompModel.objects.filter(Q(jam='00:00 WIB')).values_list('slmon', flat=True))
    # y_blanks_00 = list(ChecklistSeiscompModel.objects.filter(Q(jam='00:00 WIB')).values_list('blanks', flat=True))

    # layout00=go.Layout(title="Grafik Slmon vs Visual Monitoring Pukul 00:00 WIB", title_x=0.5, xaxis={'title':'Waktu'}, yaxis={'title':'Jumlah'})
    
    # y_blanks_00_len = []
    # for data in y_blanks_00:
    #     if data:
    #         y = len(ast.literal_eval(data))
    #         y_blanks_00_len.append(y)
    #     else:
    #         y_blanks_00_len.append(0)

    # fig00 = go.Figure(data=[
    #         go.Line(x=x_datetime, y=y_slmon_00,
    #             name='slmon 00:00 WIB',
    #             opacity=0.8, marker_color='red'),
    #         go.Line(x=x_datetime, y=y_blanks_00_len,
    #             name='blanks 00:00 WIB',
    #             opacity=0.8, marker_color='blue'),
    #         ],
    #         layout=layout00)
    
    # jam00 = plot(fig00,
    #            output_type='div', include_plotlyjs=False, show_link=False
    #            )
    
    jam00 = plot_slmon_vm(days = 100)

    # percentage of on and off of last record
    last_record = ChecklistSeiscompModel.objects.order_by('-tanggal')[:1]
    last_slmon = last_record.values('slmon')[0]['slmon'] or 0
    last_blanks = last_record.values('blanks')[0]['blanks']
    if last_blanks:
        last_blanks = len(ast.literal_eval(last_blanks))
    else:
        last_blanks = 0
    
    last_gaps = last_record.values('gaps')[0]['gaps']
    if last_gaps:
        last_gaps = len(ast.literal_eval(last_gaps))
    else:
        last_gaps = 0
    
    last_spikes = last_record.values('spikes')[0]['spikes']
    if last_spikes:
        last_spikes = len(ast.literal_eval(last_spikes))
    else:
        last_spikes = 0
    
    context = {'jam12': jam12,
               'jam00': jam00,
               'last_gaps': last_gaps,
               'last_spikes': last_spikes,
               'last_blanks': last_blanks,
               'last_slmon': last_slmon,
               'percent_gaps': round(last_gaps / station_count * 100, 2),
               'percent_spikes': round(last_spikes / station_count * 100, 2),
               'percent_blanks': round(last_blanks / station_count * 100, 2),
               'percent_slmon': round(last_slmon / station_count * 100, 2),
               'last_data': last_data
               }
    
    return render(request, "statistic_view.html", context=context)

def get_last_x_days_of_data(model, time, days: int, col: str):
    """Gets the last x days of data from the database."""
    from django.db.models import Q
    import datetime

    today = datetime.date.today()
    x_days_ago = today - datetime.timedelta(days=days)
    print(x_days_ago)
    result = list(model.objects.filter(
      tanggal__gte=x_days_ago, tanggal__lte=today).filter(
          Q(jam=time)).values_list(col, flat=True)) 
    print(result)
    return result


def plot_slmon_vm(days, time='00:00 WIB'):
    from django.db.models import Q
    from plotly.offline import plot
    import plotly.graph_objs as go
    from datetime import datetime
    import ast
    
    x_data = get_last_x_days_of_data(ChecklistSeiscompModel, time, days, 'tanggal')
    print(x_data)
    x_datetime = []
    for x in x_data:
        x = datetime.strptime(x.strftime('%Y-%m-%d')+' 00:00', '%Y-%m-%d %H:%M')
        x_datetime.append(x)

    # graphic of visual monitoring vs slmon 00:00 WIB
    y_slmon_00 = get_last_x_days_of_data(ChecklistSeiscompModel, time, days, 'slmon')
    y_blanks_00 = get_last_x_days_of_data(ChecklistSeiscompModel, time, days, 'blanks')

    layout00=go.Layout(title="Grafik Slmon vs Visual Monitoring Pukul 00:00 WIB", title_x=0.5, xaxis={'title':'Waktu'}, yaxis={'title':'Jumlah'})
    
    y_blanks_00_len = []
    for data in y_blanks_00:
        if data:
            y = len(ast.literal_eval(data))
            y_blanks_00_len.append(y)
        else:
            y_blanks_00_len.append(0)

    fig00 = go.Figure(data=[
            go.Line(x=x_datetime, y=y_slmon_00,
                name='slmon 00:00 WIB',
                opacity=0.8, marker_color='red'),
            go.Line(x=x_datetime, y=y_blanks_00_len,
                name='blanks 00:00 WIB',
                opacity=0.8, marker_color='blue'),
            ],
            layout=layout00)
    
    jam00 = plot(fig00,
               output_type='div', include_plotlyjs=False, show_link=False
               )
    return jam00


def detail_view(request, id):
    data = ChecklistSeiscompModel.objects.filter(id=id)
    return render(request, 'detail_view.html', {'data': data})