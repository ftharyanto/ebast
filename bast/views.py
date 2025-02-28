from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import BastRecordModel
from core.models import Kelompok
from .forms import BastRecordForm
from django.shortcuts import render
import requests, openpyxl, datetime, os
import pandas as pd
from django.http import JsonResponse, HttpResponse
from core.models import Operator
from io import StringIO
from openpyxl.utils.dataframe import dataframe_to_rows
from django.views import View
from django.shortcuts import redirect

class BastRecordListView(ListView):
    model = BastRecordModel
    template_name = 'bast/bastrecord_list.html'
    context_object_name = 'bastrecords'

class BastRecordCreateView(CreateView):
    model = BastRecordModel
    form_class = BastRecordForm
    template_name = 'bast/bastrecord_form.html'
    success_url = reverse_lazy('bast:bastrecord_list')

class BastRecordUpdateView(UpdateView):
    model = BastRecordModel
    form_class = BastRecordForm
    template_name = 'bast/bastrecord_form.html'
    success_url = reverse_lazy('bast:bastrecord_list')

    def form_valid(self, form):
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        record = self.get_object()
        context['existing_data'] = record.events
        return context

    def get_member_data(self, **kwargs):
        context = super().get_member_data(**kwargs)
        record = self.get_object()
        context['existing_member_data'] = record.member
        return context

class BastRecordDeleteDirectView(View):
    def post(self, request, pk, *args, **kwargs):
        try:
            record = BastRecordModel.objects.get(pk=pk)
            record.delete()
            return redirect('bast:bastrecord_list')
        except BastRecordModel.DoesNotExist:
            return HttpResponse(status=404)

# Functions
def clean_index3(data, start_datetime='2024-12-11 13:00:00', end_datetime='2024-12-11 19:00:00'):
    text = data.decode('utf-8')

    lines = text.split('\n')
    processed_lines = []
    for i, line in enumerate(lines):
        if i not in [0, 1, 3]:
            line = '|'.join(part.strip() for part in line.split('|'))
            processed_lines.append(line)

    df = pd.DataFrame([x.split('|') for x in processed_lines[1:]], columns=processed_lines[0].split('|'))
    df['Origin Time (GMT)'] = pd.to_datetime(df['Origin Time (GMT)'], format='%Y-%m-%d %H:%M:%S')

    def select_data_by_datetime_range(df, start_datetime, end_datetime):
        mask = (df['Origin Time (GMT)'] >= start_datetime) & (df['Origin Time (GMT)'] <= end_datetime)
        return df.loc[mask]

    df_selected = select_data_by_datetime_range(df, start_datetime, end_datetime)

    # sort the df_selected by 'Origin Time (GMT)'
    df_selected = df_selected.sort_values(by='Origin Time (GMT)')

    # divide the 'Origin Time (GMT)' column into 'Date' and 'OT (UTC)' columns, and put them in the first two columns, remove the 'Origin Time (GMT)' column
    df_selected['Date'] = df_selected['Origin Time (GMT)'].dt.date
    df_selected['OT (UTC)'] = df_selected['Origin Time (GMT)'].dt.time
    df_selected = df_selected[['Date', 'OT (UTC)'] + [col for col in df_selected.columns if col != 'Origin Time (GMT)']]
    df_selected = df_selected.reset_index(drop=True)

    # sort the columns to be 'Date', 'OT (UTC)', 'Lat', 'Long', 'Mag', 'D(Km)', 'Phase', 'RMS', 'Az.Gap', 'Region', but first turn the respective column names into the desired ones
    df_selected = df_selected.rename(columns={'Lon': 'Long', 'Depth': 'D(Km)', 'cntP': 'Phase', 'AZgap': 'Az. Gap', 'Remarks': 'Region'})
    df_selected = df_selected[['Date', 'OT (UTC)', 'Lat', 'Long', 'D(Km)', 'Mag', 'TypeMag', 'Region']]
    df_selected = df_selected.reset_index(drop=True)

    # add numbering to the first column
    df_selected.insert(0, 'No', range(1, len(df_selected) + 1))

    # Check for duplicate columns
    df_selected = df_selected.loc[:, ~df_selected.columns.duplicated()]

    # add MMI, terkirim M>5, and terkirim M>5 columns with empty values
    df_selected['MMI'] = ''
    df_selected['disPGN'] = ''
    df_selected['disPGR'] = '' 
    
    return df_selected

def fetch_data(request, start_datetime='2024-12-11 13:00:00', end_datetime='2024-12-11 19:00:00'):
    url = "http://202.90.198.41/index3.txt"
    response = requests.get(url)

    if response.status_code == 200:
        data = clean_index3(response.content, start_datetime, end_datetime)
        csv_data = data.to_csv(index=False)
        table_data = data.to_dict(orient='records')
        return JsonResponse({'csv': csv_data, 'table_data': table_data})
    else:
        return JsonResponse({'error': 'Failed to fetch data'}, status=500)

def get_nip(request, operator_id):
    try:
        operator = Operator.objects.get(id=operator_id)
        return JsonResponse({'nip': operator.NIP})
    except Operator.DoesNotExist:
        return JsonResponse({'error': 'Operator not found'}, status=404)

def get_member_data(request, kelompok):
    try:
        kelompok = Kelompok.objects.get(name=kelompok)
        member_data = kelompok.member.values_list('name', flat=True)
        return JsonResponse({'member_data': list(member_data)})
    except Kelompok.DoesNotExist:
        return JsonResponse({'error': 'Kelompok not found'}, status=404)

def export_to_excel(request, record_id):
    from qc.views import format_date_indonesian, get_hari_indonesia

    try:
        record = BastRecordModel.objects.get(id=record_id)
    except BastRecordModel.DoesNotExist:
        return HttpResponse(status=404)

    file_path = os.path.join(os.path.dirname(__file__), 'static/bast/BAST.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'BAST'

    tanggal = format_date_indonesian(record.bast_id[5:-2])
    hari = get_hari_indonesia(record.bast_id[5:-2])
    sheet['J4'] = f'{convert_to_roman(record.kelompok)} ({convert_to_indonesian(record.kelompok)})'
    sheet['J6'] = f'{convert_to_roman(record.kel_berikut)} ({convert_to_indonesian(record.kel_berikut)})'
    sheet['N4'] = f': {tanggal}' 
    sheet['N5'] = f': {hari}'
    sheet['N6'] = f': {record.waktu_pelaksaan}'
    # sheet['m6'] = f': {record.jam_pelaksanaan.strftime("%H:%M")} - selesai'
    sheet['G21'] = f'{record.event_indonesia}'
    sheet['G22'] = f'{record.event_luar}'
    sheet['L21'] = f': {record.event_dirasakan} event'
    sheet['L22'] = f': {record.event_dikirim} event'
    sheet['E33'] = f'IA (505) : Gaps = {record.count_gaps} ; Spike = {record.count_spikes} ; Blank = {record.count_blanks}'
    sheet['E37'] = f'{record.pulsa_poco}'
    sheet['E39'] = f'{record.poco_exp.strftime("%d %b %Y")}'
    sheet['G37'] = f'{record.pulsa_samsung}'
    sheet['G39'] = f'{record.samsung_exp.strftime("%d %b %Y")}'
    sheet['C46'] = f'Jakarta, {tanggal}'
    sheet['C54'] = f'{record.spv}'

    # import the events from the record using pandas
    events = pd.read_csv(StringIO(record.events))

    # add rows to the sheet
    rows_to_add = len(events)
    sheet.insert_rows(28, amount=rows_to_add)
    events = dataframe_to_rows(events, index=False, header=False)
    
    # insert the events to the sheet
    for r_idx, row in enumerate(events, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx+27, column=c_idx+2, value=value).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            # set the border of the first column to the left and the last column to the right, to thick
            sheet.cell(row=r_idx+27, column=2).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='medium'))
            sheet.cell(row=r_idx+27, column=17).border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='medium'))
    
    # set the inserted cell border expanded to column 17 to thin
    for r_idx in range(rows_to_add):
        for c_idx in range(14):  # Iterate up to column 17 (index 14)
            cell = sheet.cell(row=r_idx + 28, column=c_idx + 3) # Get the cell object

            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

    # Save the workbook to a BytesIO object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={record.bast_id}.xlsx'
    workbook.save(response)

    return response

def export_to_pdf(request, record_id):
    from qc.views import format_date_indonesian, get_hari_indonesia

    try:
        record = BastRecordModel.objects.get(id=record_id)
    except BastRecordModel.DoesNotExist:
        return HttpResponse(status=404)

    file_path = os.path.join(os.path.dirname(__file__), 'static/bast/BAST.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    sheet.title = 'BAST'

    tanggal = format_date_indonesian(record.bast_id[5:-2])
    hari = get_hari_indonesia(record.bast_id[5:-2])
    sheet['J4'] = f'{convert_to_roman(record.kelompok)} ({convert_to_indonesian(record.kelompok)})'
    sheet['J6'] = f'{convert_to_roman(record.kel_berikut)} ({convert_to_indonesian(record.kel_berikut)})'
    sheet['N4'] = f': {tanggal}' 
    sheet['N5'] = f': {hari}'
    sheet['N6'] = f': {record.waktu_pelaksaan}'
    # sheet['m6'] = f': {record.jam_pelaksanaan.strftime("%H:%M")} - selesai'
    sheet['G21'] = f'{record.event_indonesia}'
    sheet['G22'] = f'{record.event_luar}'
    sheet['L21'] = f': {record.event_dirasakan} event'
    sheet['L22'] = f': {record.event_dikirim} event'
    sheet['E33'] = f'IA (505) : Gaps = {record.count_gaps} ; Spike = {record.count_spikes} ; Blank = {record.count_blanks}'
    sheet['E37'] = f'{record.pulsa_poco}'
    sheet['E39'] = f'{record.poco_exp.strftime("%d %b %Y")}'
    sheet['G37'] = f'{record.pulsa_samsung}'
    sheet['G39'] = f'{record.samsung_exp.strftime("%d %b %Y")}'
    sheet['C46'] = f'Jakarta, {tanggal}'
    sheet['C54'] = f'{record.spv}'

    # import the events from the record using pandas
    events = pd.read_csv(StringIO(record.events))

    # add rows to the sheet
    rows_to_add = len(events)
    sheet.insert_rows(28, amount=rows_to_add)
    events = dataframe_to_rows(events, index=False, header=False)
    
    # insert the events to the sheet
    for r_idx, row in enumerate(events, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx+27, column=c_idx+2, value=value).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            # set the border of the first column to the left and the last column to the right, to thick
            sheet.cell(row=r_idx+27, column=2).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='medium'))
            sheet.cell(row=r_idx+27, column=17).border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='medium'))
    
    # set the inserted cell border expanded to column 17 to thin
    for r_idx in range(rows_to_add):
        for c_idx in range(14):  # Iterate up to column 17 (index 14)
            cell = sheet.cell(row=r_idx + 28, column=c_idx + 3) # Get the cell object

            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

    # temporarily save the workbook to a file
    temp_xlsx = os.path.join(os.path.dirname(__file__), f'static/bast/{record.bast_id}.xlsx')
    workbook.save(temp_xlsx)
    temp_pdf_dir = os.path.join(os.path.dirname(__file__), 'static/bast')
    temp_pdf = os.path.join(temp_pdf_dir, f'{record.bast_id}.pdf')

    import subprocess
    try:
        command = ['soffice', '--headless', '--convert-to', 'pdf:calc_pdf_Export', temp_xlsx, '--outdir', temp_pdf_dir]
        subprocess.run(command, check=True)
        print(temp_xlsx)
    except subprocess.CalledProcessError as e:
        print(f"Error converting {temp_xlsx} to PDF: {e}")
    finally:
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)

    # Read the generated PDF file and return it in the response
    with open(temp_pdf, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename={record.bast_id}.pdf'

    if os.path.exists(temp_pdf):
        os.remove(temp_pdf)

    return response

def convert_to_roman(number):
    number = int(number)
    val = [
        1000, 900, 500, 400,
        100, 90, 50, 40,
        10, 9, 5, 4,
        1
        ]
    syb = [
        "M", "CM", "D", "CD",
        "C", "XC", "L", "XL",
        "X", "IX", "V", "IV",
        "I"
        ]
    roman_num = ''
    i = 0
    while  number > 0:
        for _ in range(number // val[i]):
            roman_num += syb[i]
            number -= val[i]
        i += 1
    return roman_num

def convert_to_indonesian(number):
    number = int(number)
    indonesian_numbers = ["Nol", "Satu", "Dua", "Tiga", "Empat", "Lima", "Enam", "Tujuh", "Delapan", "Sembilan", "Sepuluh"]
    if 0 <= number < len(indonesian_numbers):
        return indonesian_numbers[number]
    else:
        return str(number)