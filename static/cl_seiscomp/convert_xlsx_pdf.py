from openpyxl import load_workbook
import xlsxwriter

# Load the Excel file
workbook = load_workbook('cl_seiscomp.xlsx')

# Create a new Excel file using xlsxwriter
workbook_xlsx = xlsxwriter.Workbook('temp.xlsx')

# Close the xlsxwriter workbook
workbook_xlsx.close()

# Use pyexcel-ods to convert the XLSX file to ODS
from pyexcel_ods import get_data, save_data
data = get_data('temp.xlsx')
save_data('your_ods_file.ods', data)